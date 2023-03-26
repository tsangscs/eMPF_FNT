from datetime import datetime
import traceback
from openpyxl import load_workbook

class Func:
    functionID = ""
    dashboardModule = ""
    busProcess = ""
    timeBox = ""
    showAndTellStatus = ""
    numPassed = 0

    def __init__(self, functionID, dashboardModule, busProcess, timeBox, showAndTellStatus):
        self.functionID = functionID
        self.dashboardModule = dashboardModule
        self.busProcess = busProcess
        self.timeBox = timeBox
        self.showAndTellStatus = showAndTellStatus

    def print(self):
        print(str(self.functionID) + ": " + str(self.dashboardModule) + ", " +
              str(self.busProcess) + ", " + str(self.timeBox) + ", " +
              str(self.showAndTellStatus) + ", " + str(self.numPassed))


class ReleaseStatus:
    sheetName = "Release Status"
    colFunctionID = 1
    colDashboardModule = 3
    colBusProcess = 4
    colTimeBox = 6
    colNumPassed = 18
    colShowAndTell = 7
    funcDict = {}

    funcMapDict = {
        "FS- UF_22_Chk_bal_0 and deceased": "FS-UF_22_Chk_bal_0anddeceased",
        "FS- UF_22a_Stop_Derisking": "FS-UF_22a_Stop_Derisking",
        "FS-BPM-DC-001": "FS-BPM-DCA-001",
        "FS-BPM-DC-005": "FS-BPM-DCA-005",
        "FS-BPM-ENG-002": "FS-BPM-ENR-002",
        "FS-BPM-INV-002D": "FS-BPM-INV-002",
        "FS-BPM-INV-002-D": "FS-BPM-INV-002",
        "FS-FS-UF-MMB-011": "FS-UF-MMB-011",
        "FS-FS-UF-WDR-009": "FS-UF-WDR-009",
        "FS-FS-UF-WDR-096": "FS-UF-WDR-096",
        "FS-FS-UF-WDR-207": "FS-UF-WDR-207",
        "FS-UD-RPT-PD-014": "FS-UF-RPT-PD-014",
        "FS-UF_06_Chk_DIS fund": "FS-UF_06_Chk_DISfund",
        "FS-UF_13_ Add_FS_instr": "FS-UF_13_Add_FS_instr",
        "FS-UF_15_De-risking existing_asset": "FS-UF_15_De-riskingexisting_asset",
        "FS-UF-CCS-Voice-Screen-QM-SA-0001": "UF-CCS-Voice-Screen-QM-SA-0001",
        "FS-UF-CCS-Voice-Screen-QM-SA-0002": "UF-CCS-Voice-Screen-QM-SA-0002",
        "FS-UF-CCS-Voice-Screen-QM-SA-0003": "UF-CCS-Voice-Screen-QM-SA-0003",
        "FS-UF-CCS-Voice-Screen-QM-SA-0004": "UF-CCS-Voice-Screen-QM-SA-0004",
        "FS-UF-CCS-Voice-Screen-QM-SA-0005": "UF-CCS-Voice-Screen-QM-SA-0005",
        "FS-UF-CON-CEE-906": "FS-UF-CON-CEE-006",
        "FS-UF-CON-EE-012": "FS-UF-CON-CEE-012",
        "FS-UF-CON-REE-058 FS-UF-CON-CEE-009": "FS-UF-CON-REE-058, FS-UF-CON-CEE-009",
        "FS-UF-CON-SVCSVC-002": "FS-UF-CON-SVCTVC-002",
        "FS-UF-DCR-MPFA-007a": "FS-UF-DCR-MPFA-007",
        "FS-UF-DM-036.": "FS-UF-DM-036",
        "FS-UF-ENR-CEE-029 FS-UF-ENR-CEE-030": "FS-UF-ENR-CEE-029, FS-UF-ENR-CEE-030",
        "FS-UF-ENR-EE-033": "FS-UF-ENR-CEE-033",
        "FS-UF-FE-ENQUIRY-MEM-005": "FS-UF-FE-EE-ENR-MEM-005",
        "FS-UF-FE-ENQUIRY-MEM-006": "FS-UF-FE-EE-ENR-MEM-006",
        "FS-UF-FE-PUB-WEB-006FS-UF-FE-3rdPARTY-004a": "FS-UF-FE-PUB-WEB-006, FS-UF-FE-3rdPARTY-004a",
        "FS-UF-FE-TRS-TVC-MEM-006": "FS-UF-FE-TRS-TVC-MEM-06",
        "FS-UF-FE-UB-MEM-007 FS-UF-WDR-PW-091": "FS-UF-FE-UB-MEM-007, FS-UF-WDR-PW-091",
        "FS-UF-FPD-001": "FS-UF-SWH-FPD-001",
        "FS-UF-INC-001": "FS-UF-INF-001",
        "FS-UF-INC-002": "FS-UF-INF-002",
        "FS-UF-INC-003": "FS-UF-INF-003",
        "FS-UF-INC-004": "FS-UF-INF-004",
        "FS-UF-INC-005": "FS-UF-INF-005",
        "FS-UF-INC-006": "FS-UF-INF-006",
        "FS-UF-INC-007": "FS-UF-INF-007",
        "FS-UF-INC-008": "FS-UF-INF-008",
        "FS-UF-INC-009": "FS-UF-INF-009",
        "FS-UF-INC-010": "FS-UF-INF-010",
        "FS-UF-INC-011": "FS-UF-INF-011",
        "FS-UF-INC-012": "FS-UF-INF-012",
        "FS-UF-INF-010D-013": "FS-UF-INF-010",
        "Fs-UF-LSP-043": "FS-UF-LSP-043",
        "FS-UF-MLFUB-UA-001-a1": "FS-UF-MLFUB-UA-001",
        "FS-UF-MMB-012 UF-UC-001": "FS-UF-MMB-012, UF-UC-001",
        "FS-UF-RPR-ARI-005": "FS-UF-RPT-ARI-005",
        "FS-UF-RPT-P": "FS-UF-RPT-PD-013",
        "FS-UF-RTP-PAR-001": "FS-UF-RPT-PAR-001",
        "FS-UF-WDR-008 UF-UC-001": "FS-UF-WDR-008, UF-UC-001",
        "FS-UF-WDR-BK-016 UF-UC-001": "FS-UF-WDR-BK-016, UF-UC-001",
        "FS-UF-WDR-PW092": "FS-UF-WDR-PW-092",
        "FS-UF-WRD-PW-008": "FS-UF-WDR-PW-008",
        "FS-US-RPT-002": "FS-UF-RPT-002",
        "MLFUB-RP-003": "FS-UF-MLFUB-RP-003",
        "S-UF-CCS-IVRS-0014": "FS-UF-CCS-IVRS-0014",
        "UF-AR-001": "FS-UF-RPT-AR-001",
        "UF-AR-002": "FS-UF-RPT-AR-002",
        "UF-DMS-001": "FS-UF-EDMS-001",
        "UF-DMS-002": "FS-UF-EDMS-002",
        "UF-DMS-003": "FS-UF-EDMS-003",
        "UF-DMS-004": "FS-UF-EDMS-004",
        "UF-DMS-005": "FS-UF-EDMS-005",
        "UF-DMS-006": "FS-UF-EDMS-006",
        "UF-DMS-007": "FS-UF-EDMS-007",
        "UF-DMS-008": "FS-UF-EDMS-008",
        "UF-DMS-009": "FS-UF-EDMS-009",
        "UF-DMS-010": "FS-UF-EDMS-010",
        "UF-DMS-011": "FS-UF-EDMS-011",
        "UF-DMS-012": "FS-UF-EDMS-012",
        "UF-DMS-013": "FS-UF-EDMS-013",
        "UF-DMS-014": "FS-UF-EDMS-014",
        "UF-ENR-CEE-008": "FS-UF-ENR-CEE-008",
        "UF-ENR-CEE-036": "FS-UF-ENR-CEE-036",
        "UF-REG-ER-013": "FS-UF-REG-ER-013",
        "UF-REG-ER-014": "FS-UF-REG-ER-014",
        "UF-REG-ER-015": "FS-UF-REG-ER-015",
        "UF-UBE-036": "FS-UF-UBE-036",
        "UF-WDR-100": "FS-UF-WDR-100",
        "UF-WDR-207": "FS-UF-WDR-207",
        "UF-WDR-PW-H14": "FS-UF-WDR-PW-H14",
        "UF-WDR-WO-008": "FS-UF-WDR-WO-008",
        "UF-WDR-WO-020": "FS-UF-WDR-WO-020"
        }

    def loadFuncDict(self, workbook):
        try:
            sheet = workbook[self.sheetName]
            print(datetime.now().strftime("%H:%M:%S") + ": Processing [" + self.sheetName + "] row 2 to max row " + str(sheet.max_row))

            for rownum in range(2, sheet.max_row+1):
                funcID = sheet.cell(row=rownum, column=self.colFunctionID).value
                print("Loading row["+str(rownum)+"]: funcID = "+ funcID)
                if funcID is not None:
                    funcID = funcID.strip()
                else:
                    continue # Blank row
                dashboardModule = sheet.cell(row=rownum, column=self.colDashboardModule).value
                if dashboardModule is not None:
                    dashboardModule = dashboardModule.strip()
                else:
                    raise Exception("Blank module for " + funcID)
                busProcess = sheet.cell(row=rownum, column=self.colBusProcess).value
                if busProcess is not None:
                    busProcess = busProcess.strip()
                else:
                    raise Exception("Blank business process for " + funcID)
                timebox = sheet.cell(row=rownum, column=self.colTimeBox).value
                if timebox is not None:
                    timebox = timebox.strip()
                else:
                    raise Exception("Blank timebox for " + funcID)
                showAndTell = sheet.cell(row=rownum, column=self.colShowAndTell).value
                if showAndTell is not None:
                    if showAndTell == "":
                        showAndTell = "Not Completed"
                    else:
                        showAndTell = showAndTell.strip()
                else:
                    showAndTell = "Not Completed"

                func = Func(funcID, dashboardModule, busProcess, timebox, showAndTell)
                self.funcDict[funcID] = func

        except Exception as ex:
            print(ex)
            traceback.print_exc()

    def printFuncDict(self):
        for f in self.funcDict.values():
            f.print()

    def updateReleaseStatusSheet(self, workbook):
        try:
            sheet = workbook[self.sheetName]
            print(datetime.now().strftime("%H:%M:%S") + ": Processing [" + self.sheetName + "] row 2 to max row " + str(sheet.max_row))

            for rownum in range(2, sheet.max_row+1):
                funcID = sheet.cell(row=rownum, column=self.colFunctionID).value
                if funcID is not None:
                    funcID = funcID.strip()
                    func = self.funcDict[funcID]
                    sheet.cell(row=rownum, column=self.colNumPassed).value = func.numPassed
                else:
                    continue

        except Exception as ex:
            print(ex)
            traceback.print_exc()

class TestCase:
    testName = ""
    dashboardModule = ""
    busProcess = ""
    timeBox = ""
    passStatus = ""
    showAndTellStatus = ""
    funclist = {}
    error = ""

    def __init__(self, testName):
        self.testName = testName

    def print(self):
        print(self.testName + ", " +
              self.dashboardModule + ", " +
              self.busProcess + ", " +
              self.timeBox + ", " +
              self.passStatus + ", " +
              self.showAndTellStatus + ", " +
              ":".join(self.funclist)
              )

    def appendError(self, errMsg):
        if self.error != "":
            self.error += ": "
        self.error += errMsg

class TestCaseSheet:
    sheetName = "Main Sheet"
    colTestName = 1
    colFNTStatus = 3
    colModule = 4
    colBusProcess = 6
    colRQID = 7
    colShowAndTellStatus = 8
    colTimeBox = 21
    colErrMsg = 22

    def updateTestCaseSheet(self, workbook):
        try:
            sheet = workbook[self.sheetName]
            sheet.cell(row=1, column=self.colTimeBox).value = "Time Box"
            sheet.cell(row=1, column=self.colErrMsg).value = "Error Message"

            #Read Function Release Status
            releaseStatus = ReleaseStatus()
            releaseStatus.loadFuncDict(workbook)
            print(datetime.now().strftime("%H:%M:%S") + ": Processing [" + self.sheetName + "] row 2 to max row " + str(sheet.max_row))

            for rownum in range(2, sheet.max_row+1):
                testName = sheet.cell(row=rownum, column=self.colTestName).value
                if testName is None:
                    continue
                elif len(testName.strip()) == 0:
                    continue
                try:
                    testCase = TestCase(testName)

                    # Check Pass Status
                    testCase.passStatus = sheet.cell(row=rownum, column=self.colFNTStatus).value

                    # Check Functions
                    rqidlist = sheet.cell(row=rownum, column=self.colRQID).value
                    if rqidlist is not None:
                        rqidlist = rqidlist.strip()
                        if len(rqidlist) != 0 and rqidlist != "#N/A":
    #                        if rqidlist[len(rqidlist)] == ",":
    #                            rqidlist = rqidlist[0:len(rqidlist)-2]
                            rqidlist = rqidlist.replace(chr(10), ",")
                            testCase.funclist = rqidlist.split(",")

                    if testCase.funclist is None or len(testCase.funclist) == 0:
                        testCase.appendError("RQID is #N/A or blank")
                    else:
                        dashboardModule = []
                        busProcess = []
                        for rqid in testCase.funclist:
                            funcID = rqid.strip()
                            if len(funcID) == 0:
                                continue
                            func = releaseStatus.funcDict.get(funcID)
                            # Found Functions
                            if func is not None:
                                # Update Dashboard Module
                                try:
                                    dashboardModule.index(func.dashboardModule)
                                except Exception as ex:
                                    dashboardModule.append(func.dashboardModule)
                                    dashboardModule.sort()
                                    testCase.dashboardModule = ":".join(dashboardModule)

                                # Update Business Process
                                try:
                                    busProcess.index(func.busProcess)
                                except Exception as ex:
                                    busProcess.append(func.busProcess)
                                    busProcess.sort()
                                    testCase.busProcess = ":".join(busProcess)

                                # Update Test Case Timebox
                                if func.timeBox in ["Day 2", "De-scoped"]:
                                    testCase.timeBox = func.timeBox
                                elif func.timeBox == "TX13":
                                    if testCase.timeBox in ["", "TX1-10", "TX-11", "TX-12"]:
                                        testCase.timeBox = func.timeBox
                                elif func.timeBox == "TX12":
                                    if testCase.timeBox in ["", "TX1-10", "TX-11"]:
                                        testCase.timeBox = func.timeBox
                                elif func.timeBox == "TX11":
                                    if testCase.timeBox in ["", "TX1-10"]:
                                        testCase.timeBox = func.timeBox
                                elif func.timeBox == "TX1-10":
                                    if testCase.timeBox == "":
                                        testCase.timeBox = func.timeBox
                                else:
                                    testCase.appendError("Invalid timebox of " + funcID + " = " + func.timeBox)

                                # Update Test Case Show & Tell Status
                                if func.showAndTellStatus == "Not Completed":
                                    testCase.showAndTellStatus = "Not Completed"
                                elif func.showAndTellStatus in ["N/A", "Done"]:
                                    if testCase.showAndTellStatus != "Not Completed":
                                        testCase.showAndTellStatus = "Done"
                                else:
                                    testCase.appendError("Invalid S&T Status of " + funcID + " = " + func.showAndTellStatus)

                                # Update function pass status
                                if testCase.passStatus[0:6] == "Passed":
                                    func.numPassed += 1
                                    releaseStatus.funcDict.update({funcID: func})
                            else:
                                testCase.appendError(funcID + " not valid")

                        # Update test case
                        sheet.cell(row=rownum, column=self.colModule).value = testCase.dashboardModule
                        sheet.cell(row=rownum, column=self.colBusProcess).value = testCase.busProcess
                        sheet.cell(row=rownum, column=self.colTimeBox).value = testCase.timeBox
                        sheet.cell(row=rownum, column=self.colShowAndTellStatus).value = testCase.showAndTellStatus
                    # end if
                    sheet.cell(row=rownum, column=self.colErrMsg).value = testCase.error

                except Exception as ex:
                    errMsg = str(ex)
                    print(datetime.now().strftime("%H:%M:%S") +
                          ": Error found in [" + self.sheetName + "] at row " + str(rownum) + " Error: " + errMsg)

            # end for

            #print("Result Function Status: ")
            #releaseStatus.printFuncDict()
            releaseStatus.updateReleaseStatusSheet(workbook)

        except Exception as ex:
            errMsg = str(ex)
            print("Error found at row[" + str(rownum) + "]: " + errMsg)
            traceback.print_exc()
            if testCase is not None:
                if testCase.error is not None:
                    errMsg = errMsg + ", " + testCase.error
            sheet.cell(row=rownum, column=self.colErrMsg).value = testCase.error

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    infile = input("Enter filename of FNT Execution Report: ")
    extensionIdx = infile.rfind(".xlsx")
    outfile = infile[0:extensionIdx] + "_out.xlsx"
    print("Output file = " + outfile)

    print(datetime.now().strftime("%H:%M:%S") + ": Loading input workbook ...")
    workbook = load_workbook(filename=infile)

    print(datetime.now().strftime("%H:%M:%S") + ": Processing ...")
    testCaseSheet = TestCaseSheet()
    testCaseSheet.updateTestCaseSheet(workbook)

    print(datetime.now().strftime("%H:%M:%S") + ": Saving output workbook ...")
    workbook.save(outfile)
    print(datetime.now().strftime("%H:%M:%S") + ": Job completed!")
